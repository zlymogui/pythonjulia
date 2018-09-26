#!/usr/bin/python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def write_excel(filename,sheetname,row,column,output):
    wb = load_workbook(filename)
    ws = wb[sheetname]
    ws.cell(row, column).value = output
    wb.save(filename)