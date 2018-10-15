#!/usr/bin/python
# -*- coding: utf-8 -*-

from openpyxl import load_workbook

def read_excel(filename,sheetname):
    # 打开一个workbook
    wb = load_workbook(filename)
    # 获取所有表格(worksheet)的名字
    #sheets = wb.sheetnames
    # 第一个表格的名称
    #sheet_first = sheets[0]
    # 获取特定的worksheet
    ws = wb[sheetname]
    ncol = ws.max_column
    nrow = ws.max_row

    return ws,ncol,nrow
